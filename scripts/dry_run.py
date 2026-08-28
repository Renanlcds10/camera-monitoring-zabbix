#!/usr/bin/env python3

import argparse
import os
import sys

import requests
from openpyxl import load_workbook


# ============================================================
# ARGUMENTOS
# ============================================================

parser = argparse.ArgumentParser(
    description="Dry run da importação de câmeras."
)

parser.add_argument(
    "--excel",
    required=True
)

parser.add_argument(
    "--sheet",
    required=True
)

parser.add_argument(
    "--group-name",
    required=True
)

args = parser.parse_args()


# ============================================================
# CONFIGURAÇÕES
# ============================================================

ZABBIX_URL = os.environ.get("ZABBIX_URL")
TOKEN = os.environ.get("ZABBIX_TOKEN")

if not ZABBIX_URL:
    print("ERRO: variável ZABBIX_URL não encontrada.")
    sys.exit(1)

if not TOKEN:
    print("ERRO: variável ZABBIX_TOKEN não encontrada.")
    sys.exit(1)


# ============================================================
# API
# ============================================================

def api(method, params):

    response = requests.post(
        ZABBIX_URL,
        headers={
            "Content-Type": "application/json-rpc",
            "Authorization": f"Bearer {TOKEN}",
        },
        json={
            "jsonrpc": "2.0",
            "method": method,
            "params": params,
            "id": 1,
        },
        timeout=15,
    )

    response.raise_for_status()

    data = response.json()

    if "error" in data:
        raise RuntimeError(
            f"Erro Zabbix API: {data['error']}"
        )

    return data["result"]


# ============================================================
# INÍCIO
# ============================================================

print("=" * 70)
print("IMPORTADOR DE CÂMERAS - DRY RUN")
print("NENHUMA ALTERAÇÃO SERÁ FEITA NO ZABBIX")
print("=" * 70)
print()


# ============================================================
# CONSULTAR HOSTS EXISTENTES
# ============================================================

print("Consultando hosts existentes...")

hosts = api(
    "host.get",
    {
        "output": ["hostid", "host", "name"],
        "selectInterfaces": ["ip"],
    },
)

nomes_existentes = {}
ips_existentes = {}

for host in hosts:

    nomes_existentes[
        host["name"].strip().lower()
    ] = host

    for interface in host.get("interfaces", []):

        ip = interface.get("ip")

        if ip:
            ips_existentes[ip] = host


# ============================================================
# ABRIR PLANILHA
# ============================================================

print(f"Lendo planilha: {args.excel}")
print(f"Aba: {args.sheet}")
print()

workbook = load_workbook(
    args.excel,
    data_only=True
)

sheet = workbook[args.sheet]


# ============================================================
# RESULTADOS
# ============================================================

criaria = []
existentes = []
conflitos = []


# ============================================================
# VALIDAR CÂMERAS
# ============================================================

for row in sheet.iter_rows(
    min_row=2,
    values_only=True
):

    numero, nome, ip, porta, grupo = row

    if not nome or not ip:
        continue

    nome = str(nome).strip()
    ip = str(ip).strip()
    porta = int(porta)

    print("-" * 70)

    # --------------------------------------------------------
    # NOME JÁ EXISTE
    # --------------------------------------------------------

    if nome.lower() in nomes_existentes:

        print(f"[JÁ EXISTE] {nome}")
        print(f"             IP Excel: {ip}")

        existentes.append(
            (nome, ip)
        )

        continue

    # --------------------------------------------------------
    # IP JÁ ESTÁ EM USO
    # --------------------------------------------------------

    if ip in ips_existentes:

        host_existente = ips_existentes[ip]

        print(f"[IP EM USO]  {nome}")
        print(f"             IP: {ip}")
        print(
            "             Já usado por: "
            f"{host_existente['host']}"
        )

        conflitos.append(
            (
                nome,
                ip,
                host_existente["host"]
            )
        )

        continue

    # --------------------------------------------------------
    # SERIA CRIADO
    # --------------------------------------------------------

    print(f"[CRIARIA]    {nome}")
    print(f"             IP: {ip}")
    print(f"             Porta: {porta}")
    print(f"             Grupo: {args.group_name}")

    criaria.append(
        (nome, ip, porta)
    )


# ============================================================
# RESUMO
# ============================================================

print()
print("=" * 70)
print("RESUMO DO DRY RUN")
print("=" * 70)

print(f"Seriam criadas:     {len(criaria)}")
print(f"Já existem:         {len(existentes)}")
print(f"IPs em conflito:    {len(conflitos)}")

if conflitos:

    print()
    print("ATENÇÃO - IPs em conflito:")

    for nome, ip, host_existente in conflitos:

        print(
            f"  {ip} | {nome} | "
            f"já pertence a {host_existente}"
        )

print()
print("DRY RUN FINALIZADO.")
print("Nenhum host foi criado ou alterado.")
