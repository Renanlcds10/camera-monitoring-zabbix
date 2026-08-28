#!/usr/bin/env python3

import argparse
import os
import sys

import requests
from openpyxl import load_workbook


# ============================================================
# ARGUMENTOS
# ============================================================

parser = argparse.ArgumentParser(
    description="Importação de câmeras para o Zabbix."
)

parser.add_argument(
    "--excel",
    required=True,
    help="Caminho da planilha Excel."
)

parser.add_argument(
    "--sheet",
    required=True,
    help="Nome da aba da planilha."
)

parser.add_argument(
    "--group-id",
    required=True,
    help="ID do grupo de hosts no Zabbix."
)

parser.add_argument(
    "--server-tag",
    required=True,
    help="Valor usado na tag servidor."
)

parser.add_argument(
    "--template-tcp-id",
    required=True,
    help="ID do template responsável pela checagem TCP."
)

parser.add_argument(
    "--template-icmp-id",
    required=True,
    help="ID do template ICMP."
)

parser.add_argument(
    "--port-macro",
    default="{$DGUARD.PORT}",
    help="Macro usada para a porta TCP."
)

parser.add_argument(
    "--host-prefix",
    default="camera",
    help="Prefixo usado no nome técnico do host."
)

args = parser.parse_args()


# ============================================================
# CONFIGURAÇÕES
# ============================================================

ZABBIX_URL = os.environ.get("ZABBIX_URL")
TOKEN = os.environ.get("ZABBIX_TOKEN")

if not ZABBIX_URL:
    print("ERRO: variável ZABBIX_URL não encontrada.")
    sys.exit(1)

if not TOKEN:
    print("ERRO: variável ZABBIX_TOKEN não encontrada.")
    print()
    print("Execute antes:")
    print("read -s ZABBIX_TOKEN")
    print("export ZABBIX_TOKEN")
    sys.exit(1)


# ============================================================
# FUNÇÃO PARA CHAMAR A API DO ZABBIX
# ============================================================

def api(method, params):

    response = requests.post(
        ZABBIX_URL,
        headers={
            "Content-Type": "application/json-rpc",
            "Authorization": f"Bearer {TOKEN}",
        },
        json={
            "jsonrpc": "2.0",
            "method": method,
            "params": params,
            "id": 1,
        },
        timeout=20,
    )

    response.raise_for_status()

    data = response.json()

    if "error" in data:
        raise RuntimeError(
            f"Erro Zabbix API: {data['error']}"
        )

    return data["result"]


# ============================================================
# CONFIRMAÇÃO
# ============================================================

print("=" * 70)
print("IMPORTAÇÃO REAL DE CÂMERAS")
print("=" * 70)
print()

confirmacao = input(
    "Isso CRIARÁ hosts no Zabbix. Digite IMPORTAR para continuar: "
)

if confirmacao != "IMPORTAR":
    print("Importação cancelada.")
    sys.exit(0)


# ============================================================
# CONSULTAR HOSTS EXISTENTES
# ============================================================

print()
print("Consultando hosts existentes no Zabbix...")

hosts = api(
    "host.get",
    {
        "output": ["hostid", "host", "name"],
        "selectInterfaces": ["ip"],
    },
)

nomes_existentes = {}
ips_existentes = {}

for host in hosts:

    nome_visivel = host["name"].strip().lower()

    nomes_existentes[nome_visivel] = host

    for interface in host.get("interfaces", []):

        ip = interface.get("ip")

        if ip:
            ips_existentes[ip] = host

print(f"Hosts encontrados: {len(hosts)}")


# ============================================================
# ABRIR PLANILHA
# ============================================================

print()
print(f"Lendo planilha: {args.excel}")
print(f"Aba: {args.sheet}")

workbook = load_workbook(
    args.excel,
    data_only=True
)

sheet = workbook[args.sheet]


# ============================================================
# CONTADORES
# ============================================================

criados = []
ignorados = []
falhas = []


# ============================================================
# LER CÂMERAS
# ============================================================

for row in sheet.iter_rows(
    min_row=2,
    values_only=True
):

    numero, nome, ip, porta, grupo = row

    if not nome or not ip or not porta:
        continue

    nome = str(nome).strip()
    ip = str(ip).strip()
    porta = str(int(porta))

    print()
    print("-" * 70)
    print(nome)
    print(f"IP: {ip}")
    print(f"Porta: {porta}")

    # --------------------------------------------------------
    # VERIFICAR DUPLICIDADE POR NOME
    # --------------------------------------------------------

    if nome.lower() in nomes_existentes:

        host_existente = nomes_existentes[nome.lower()]

        print("[IGNORADO] Host com esse nome já existe.")
        print(f"Host ID: {host_existente['hostid']}")

        ignorados.append(
            (nome, ip, "nome existente")
        )

        continue

    # --------------------------------------------------------
    # VERIFICAR DUPLICIDADE POR IP
    # --------------------------------------------------------

    if ip in ips_existentes:

        host_existente = ips_existentes[ip]

        print(
            "[IGNORADO] IP já está em uso por: "
            f"{host_existente['host']}"
        )

        ignorados.append(
            (
                nome,
                ip,
                f"IP usado por {host_existente['host']}"
            )
        )

        continue

    # --------------------------------------------------------
    # NOME TÉCNICO DO HOST
    # --------------------------------------------------------

    host_tecnico = (
        f"{args.host_prefix}-"
        f"{ip.replace('.', '-')}"
    )

    # --------------------------------------------------------
    # PARÂMETROS DO HOST
    # --------------------------------------------------------

    params = {
        "host": host_tecnico,
        "name": nome,

        "groups": [
            {
                "groupid": args.group_id
            }
        ],

        "interfaces": [
            {
                "type": 1,
                "main": 1,
                "useip": 1,
                "ip": ip,
                "dns": "",
                "port": "10050",
            }
        ],

        "templates": [
            {
                "templateid": args.template_tcp_id
            },
            {
                "templateid": args.template_icmp_id
            },
        ],

        "tags": [
            {
                "tag": "tipo",
                "value": "camera"
            },
            {
                "tag": "sistema",
                "value": "dguard"
            },
            {
                "tag": "servidor",
                "value": args.server_tag
            },
        ],

        "macros": [
            {
                "macro": args.port_macro,
                "value": porta
            }
        ],
    }

    # --------------------------------------------------------
    # CRIAR HOST
    # --------------------------------------------------------

    try:

        resultado = api(
            "host.create",
            params
        )

        hostid = resultado["hostids"][0]

        print(
            f"[CRIADO] Host ID: {hostid}"
        )

        criados.append(
            (nome, ip, porta, hostid)
        )

        # Atualiza os caches locais para impedir
        # duplicidade durante a própria execução.

        nomes_existentes[nome.lower()] = {
            "host": host_tecnico,
            "hostid": hostid,
        }

        ips_existentes[ip] = {
            "host": host_tecnico,
            "hostid": hostid,
        }

    except Exception as erro:

        print(f"[ERRO] {erro}")

        falhas.append(
            (nome, ip, str(erro))
        )


# ============================================================
# RESULTADO
# ============================================================

print()
print("=" * 70)
print("RESULTADO DA IMPORTAÇÃO")
print("=" * 70)

print(f"Criados:    {len(criados)}")
print(f"Ignorados:  {len(ignorados)}")
print(f"Falhas:     {len(falhas)}")

if falhas:

    print()
    print("FALHAS:")

    for nome, ip, erro in falhas:

        print(
            f"- {nome} | {ip}"
        )

        print(
            f"  {erro}"
        )

print()
print("IMPORTAÇÃO FINALIZADA.")
